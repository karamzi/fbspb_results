from typing import List

from openpyxl.worksheet.worksheet import Worksheet

from models import Player, BaseResultParse


class FinalsB(BaseResultParse):
    type = 'finals_b'

    def parse_sheet(self, ws: Worksheet, target: List[Player], column: str):

        while not isinstance(ws[self.get_cell_address()].value, int):
            self.change_row()

        while ws[self.get_cell_address()].value is not None:
            count = 0
            self.change_column()
            name = ws[self.get_cell_address()].value
            player = self.find_player(target, name)

            if not player:
                player = Player()
                player.name = name
                target.append(player)

            self.change_column()
            player.games.append(ws[self.get_cell_address()].value)
            self.change_column()

            if isinstance(ws[self.get_cell_address()].value, int):
                player.games.append(ws[self.get_cell_address()].value)

            self.change_column(by=2)
            player.place = ws[self.get_cell_address()].value
            self.next_line(column=column)

            while not isinstance(ws[self.get_cell_address()].value, int) and count != 5:
                self.change_row()
                count += 1

    def parse(self):
        column = 'A'
        self.pointer = [column, 1]
        ws = self.wb[self.wb.sheetnames[0]]
        self.parse_sheet(ws, self.man, column)

        column = 'H'
        self.pointer = [column, 1]
        self.parse_sheet(ws, self.woman, column)
